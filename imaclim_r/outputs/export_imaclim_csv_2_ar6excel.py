# =============================================
# Contact: <imaclim.r.world@gmail.com>
# Licence: AGPL-3.0
# Authors:
#     Florian Leblanc
#     (CIRED - CNRS/AgroParisTech/ENPC/EHESS/CIRAD)
# =============================================

# -*- coding: utf-8 -*-
# use /home/leblanc/python2.7/bin/python2.7


import csv
import openpyxl as oxl
import numpy as np 
from shutil import copyfile
import argparse
import sys

parser = argparse.ArgumentParser()
parser.add_argument("-s", "--scy", type=str)
parser.add_argument("-f", "--filenamedata", type=str)

args = parser.parse_args()

modelname = 'IMACLIM 2.0'

scy= args.scy
filenamedata= args.filenamedata 
#scy = 'EMFBASE'
#filenamedata=sel_outputs_ar6001.tsv

# data
datamain = np.genfromtxt( filenamedata, delimiter='|')

with open('../list_outputs_str.csv', 'rb') as f:
    reader = csv.reader(f)
    outputvar = list(reader)
with open('../list_template_region.csv', 'rb') as f:
    reader = csv.reader(f)
    outputreg = list(reader)
with open('../list_outputs_units.csv', 'rb') as f:
    reader = csv.reader(f)
    outputunit = list(reader)

with open('../list_template_comments.csv', 'rb') as f:
    reader = csv.reader(f)
    outputcomment = list(reader)


list_year=list()
for elt in np.arange(2015,2101,5):
  list_year.append(elt)

nb_year=len(list_year)

filename_ref='../2022-09_08_template_NAVIGATE.xlsx'
filename='./IMACLIM_navigate_outputs_'+scy+'.xlsx'
copyfile(filename_ref, filename)

workbook = oxl.load_workbook(filename)
ws = workbook.create_sheet("data")

row=1
col=0

mainheader = ['Model','Scenario','Region','Variable','Unit']


# write header
for elt in mainheader:
    col+=1
    ws.cell(row, col).value = elt
for elt in list_year:
    col+=1
    ws.cell(row, col).value = elt

for row in range(datamain.shape[0]):
    col=1
    ws.cell(row+2, col).value = modelname
    col+=1
    ws.cell(row+2, col).value = scy
    col+=1
    ws.cell(row+2, col).value = outputreg[row][0]
    col+=1
    ws.cell(row+2, col).value = outputvar[row][0]
    col+=1
    if outputunit[row]==[]:#handle no unit
        ws.cell(row+2, col).value = ''
    else:
        ws.cell(row+2, col).value = outputunit[row][0]
    for jj in range(nb_year):
        ws.cell(row+2, col+jj+1).number_format = '0.000000000000'
        ws.cell(row+2, col+jj+1).value =  str(datamain[row, jj])


ws = workbook.create_sheet("comments")

mainheader = ['Model','Scenario','Region','Variable','Year','Comment']

row=1
col=0
# write header
for elt in mainheader:
    col+=1
    ws.cell(row, col).value = elt

for row in range( len(outputcomment)):
    col=1
    ws.cell(row+2, col).value = modelname
    col+=1
    ws.cell(row+2, col).value = scy
    col+=1
    ws.cell(row+2, col).value = ''
    col+=1
    ws.cell(row+2, col).value = outputvar[row][0]
    col+=1
    ws.cell(row+2, col).value = ''
    col+=1
    if outputcomment[row]==[]:
        ws.cell(row+2, col).value = ''
    else:
        strtpt = ''
        for elt in outputcomment[row]:
            strtpt = strtpt + elt
        ws.cell(row+2, col).value = strtpt

workbook.save(filename)
